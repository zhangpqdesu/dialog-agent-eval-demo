#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = Path(
    "/Users/zhangpq/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_jeyv1v60ynqc22_6797/msg/file/2026-04/"
    "命题二：外呼任务对话模型指令示例.xlsx"
)
SCHEMA_PATH = ROOT / "data" / "schemas" / "dialog_instruction.schema.json"
OUTPUT_PATH = ROOT / "data" / "processed" / "dialog_instruction_examples.json"
EVAL_SCHEMA_PATH = ROOT / "data" / "schemas" / "dialog_instruction_eval.schema.json"
EVAL_OUTPUT_PATH = ROOT / "data" / "processed" / "dialog_instruction_eval_examples.json"


HEADING_ALIASES = {
    "role": "role",
    "task": "task",
    "opening line": "opening_line",
    "constraints": "constraints",
    "call flow": "call_flow",
    "conversation flow": "call_flow",
    "knowledge points (faq)": "knowledge_points",
    "knowledge points": "knowledge_points",
    "faq": "knowledge_points",
}
CANONICAL_KEYS = set(HEADING_ALIASES.values())


@dataclass
class SectionNode:
    level: int
    title: str
    body_lines: list[str] = field(default_factory=list)
    children: list["SectionNode"] = field(default_factory=list)


def ensure_dirs() -> None:
    SCHEMA_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    EVAL_SCHEMA_PATH.parent.mkdir(parents=True, exist_ok=True)
    EVAL_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)


def clean_text(value: str) -> str:
    text = value.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    return text.strip()


def clean_inline_text(value: str) -> str:
    text = clean_text(value)
    return re.sub(r"\s+", " ", text)


def normalize_heading(raw: str) -> str:
    title = clean_inline_text(raw).rstrip(":")
    return HEADING_ALIASES.get(title.lower(), title.lower().replace(" ", "_"))


def parse_section_tree(text: str) -> list[SectionNode]:
    roots: list[SectionNode] = []
    stack: list[SectionNode] = []

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        heading_match = re.match(r"^(#{1,6})\s*(.+?)\s*$", line)
        if heading_match:
            level = len(heading_match.group(1))
            heading_text = heading_match.group(2).strip()
            inline_body = ""

            if ":" in heading_text:
                possible_title, possible_body = heading_text.split(":", 1)
                if possible_title.strip().lower() in HEADING_ALIASES:
                    heading_text = possible_title.strip()
                    inline_body = possible_body.strip()

            node = SectionNode(level=level, title=heading_text)
            if inline_body:
                node.body_lines.append(inline_body)

            while stack and stack[-1].level >= level:
                stack.pop()

            if stack:
                stack[-1].children.append(node)
            else:
                roots.append(node)

            stack.append(node)
            continue

        if stack:
            stack[-1].body_lines.append(line)

    return roots


def walk_nodes(nodes: list[SectionNode]) -> list[SectionNode]:
    walked: list[SectionNode] = []
    for node in nodes:
        walked.append(node)
        walked.extend(walk_nodes(node.children))
    return walked


def extract_canonical_sections(text: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current_key: str | None = None

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        heading_match = re.match(r"^(#{1,6})\s*(.+?)\s*$", line)
        if heading_match:
            heading_text = heading_match.group(2).strip()
            inline_body = ""

            if ":" in heading_text:
                possible_title, possible_body = heading_text.split(":", 1)
                normalized_title = normalize_heading(possible_title)
                if normalized_title in CANONICAL_KEYS:
                    current_key = normalized_title
                    sections.setdefault(current_key, [])
                    inline_body = possible_body.strip()
                    if inline_body:
                        sections[current_key].append(inline_body)
                    continue

            normalized_heading = normalize_heading(heading_text)
            if normalized_heading in CANONICAL_KEYS:
                current_key = normalized_heading
                sections.setdefault(current_key, [])
                continue

        if current_key is not None:
            sections[current_key].append(line)

    return {key: clean_text("\n".join(lines)) for key, lines in sections.items()}


def split_bullets(text: str) -> list[str]:
    items: list[str] = []
    current: list[str] = []

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if re.match(r"^(-|\*|\d+\.)\s+", line):
            if current:
                items.append(clean_inline_text(" ".join(current)))
            current = [re.sub(r"^(-|\*|\d+\.)\s+", "", line)]
        else:
            if current:
                current.append(line)
            else:
                current = [line]

    if current:
        items.append(clean_inline_text(" ".join(current)))

    return items


def section_body(node: SectionNode) -> str:
    return clean_text("\n".join(node.body_lines))


def extract_special_rules(constraints: list[str]) -> dict[str, Any]:
    rules: dict[str, Any] = {
        "forbidden_expressions": [],
    }

    for item in constraints:
        max_chars = re.search(r"(\d+)\s*[-~至]?\s*(\d+)?\s*个字", item)
        if max_chars:
            lower = int(max_chars.group(1))
            upper = int(max_chars.group(2) or max_chars.group(1))
            rules["max_chars_per_turn"] = upper
            rules["suggested_char_range"] = [lower, upper]

        if "不说" in item:
            forbidden = re.findall(r'"([^"]+)"', item)
            if forbidden:
                rules["forbidden_expressions"].extend(forbidden)

        if "在开车" in item and "稍后再打" in item:
            rules["driving_rule"] = "user_driving_then_politely_end_call"

        if "超出职责范围" in item and "回电" in item:
            quoted = re.findall(r'"([^"]+)"', item)
            if quoted:
                rules["out_of_scope_reply"] = quoted[0]

        if "挂断" in item:
            rules["has_end_call_condition"] = True

    if not rules["forbidden_expressions"]:
        rules.pop("forbidden_expressions")

    return rules


def infer_domain(example: dict[str, Any]) -> str:
    role_and_task = f"{example['role']} {example['task']}"
    if any(keyword in role_and_task for keyword in ["骑手", "站长", "飞毛腿", "配送"]):
        return "rider_operations"
    if any(keyword in role_and_task for keyword in ["课程", "直播", "商家", "培训机构"]):
        return "merchant_education_saas"
    return "general_outbound_service"


def collect_required_information(example: dict[str, Any]) -> list[str]:
    items: list[str] = []

    for step in example["call_flow"]["steps"]:
        instruction = clean_inline_text(step["instruction"])
        if instruction:
            items.append(instruction)
        for substep in step.get("substeps", []):
            sub_instruction = clean_inline_text(substep["instruction"])
            if sub_instruction:
                items.append(sub_instruction)

    for point in example["knowledge_points"]:
        items.append(clean_inline_text(point))

    deduped: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item and item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped


def build_success_criteria(example: dict[str, Any]) -> list[dict[str, Any]]:
    criteria: list[dict[str, Any]] = [
        {
            "criterion_id": "goal_delivery",
            "category": "task_completion",
            "description": example["task"],
            "priority": "high",
            "evidence_type": "semantic_match",
            "source_fields": ["task"],
        },
        {
            "criterion_id": "opening_line_used",
            "category": "flow_adherence",
            "description": "对话开头应完成规定的身份确认或开场告知。",
            "priority": "medium",
            "evidence_type": "semantic_match",
            "source_fields": ["opening_line"],
        },
    ]

    for idx, step in enumerate(example["call_flow"]["steps"], start=1):
        criteria.append(
            {
                "criterion_id": f"flow_step_{idx}",
                "category": "flow_adherence",
                "description": clean_inline_text(step["instruction"] or step["title"]),
                "priority": "high" if idx <= 2 else "medium",
                "evidence_type": "step_coverage",
                "source_fields": ["call_flow.steps"],
            }
        )
        for sub_idx, substep in enumerate(step.get("substeps", []), start=1):
            criteria.append(
                {
                    "criterion_id": f"flow_step_{idx}_{sub_idx}",
                    "category": "flow_adherence",
                    "description": clean_inline_text(substep["instruction"] or substep["title"]),
                    "priority": "medium",
                    "evidence_type": "step_coverage",
                    "source_fields": ["call_flow.steps.substeps"],
                }
            )

    if example["knowledge_points"]:
        criteria.append(
            {
                "criterion_id": "faq_consistency",
                "category": "knowledge_accuracy",
                "description": "当用户追问时，回答应与给定知识点一致，不得擅自编造规则。",
                "priority": "high",
                "evidence_type": "knowledge_grounding",
                "source_fields": ["knowledge_points"],
            }
        )

    constraints = example["constraints"]
    if "max_chars_per_turn" in constraints:
        criteria.append(
            {
                "criterion_id": "turn_length_limit",
                "category": "constraint_following",
                "description": f"单轮回复长度不应超过 {constraints['max_chars_per_turn']} 个字。",
                "priority": "high",
                "evidence_type": "rule_match",
                "source_fields": ["constraints.max_chars_per_turn"],
            }
        )
    if constraints.get("forbidden_expressions"):
        criteria.append(
            {
                "criterion_id": "forbidden_expressions",
                "category": "constraint_following",
                "description": "回复中不得出现禁用表达。",
                "priority": "high",
                "evidence_type": "string_match",
                "source_fields": ["constraints.forbidden_expressions"],
            }
        )

    return criteria


def build_failure_conditions(example: dict[str, Any]) -> list[dict[str, Any]]:
    failure_conditions: list[dict[str, Any]] = [
        {
            "condition_id": "miss_primary_task",
            "severity": "high",
            "description": "未完成任务目标，核心通知或引导信息没有传达给用户。",
            "trigger_type": "missing_goal",
            "source_fields": ["task"],
        },
        {
            "condition_id": "miss_required_flow",
            "severity": "high",
            "description": "遗漏关键流程步骤，导致对话未按任务脚本推进。",
            "trigger_type": "missing_step",
            "source_fields": ["call_flow.steps"],
        },
    ]

    constraints = example["constraints"]
    if "max_chars_per_turn" in constraints:
        failure_conditions.append(
            {
                "condition_id": "exceed_turn_length",
                "severity": "medium",
                "description": f"任一轮回复字数超过 {constraints['max_chars_per_turn']} 个字。",
                "trigger_type": "constraint_violation",
                "source_fields": ["constraints.max_chars_per_turn"],
            }
        )
    if constraints.get("forbidden_expressions"):
        failure_conditions.append(
            {
                "condition_id": "use_forbidden_expression",
                "severity": "medium",
                "description": "使用了禁用口头语或禁用表达。",
                "trigger_type": "constraint_violation",
                "source_fields": ["constraints.forbidden_expressions"],
            }
        )
    if "out_of_scope_reply" in constraints:
        failure_conditions.append(
            {
                "condition_id": "bad_out_of_scope_handling",
                "severity": "medium",
                "description": "遇到超出职责范围的问题时，没有按规定兜底回复。",
                "trigger_type": "exception_handling_failure",
                "source_fields": ["constraints.out_of_scope_reply"],
            }
        )
    if constraints.get("driving_rule"):
        failure_conditions.append(
            {
                "condition_id": "continue_when_user_driving",
                "severity": "high",
                "description": "用户表明在开车后，仍继续讲解而未结束通话。",
                "trigger_type": "exception_handling_failure",
                "source_fields": ["constraints.driving_rule"],
            }
        )
    if constraints.get("has_end_call_condition"):
        failure_conditions.append(
            {
                "condition_id": "miss_end_call_condition",
                "severity": "medium",
                "description": "出现明确挂断条件时，未按要求结束通话。",
                "trigger_type": "exception_handling_failure",
                "source_fields": ["constraints.items"],
            }
        )

    return failure_conditions


def build_user_simulator_profiles(example: dict[str, Any]) -> list[dict[str, Any]]:
    domain = infer_domain(example)
    profiles: list[dict[str, Any]] = [
        {
            "profile_id": "cooperative_user",
            "persona": "配合型用户",
            "goal": "愿意听完说明，并就关键点做少量确认。",
            "style": ["简短回应", "接受引导", "会确认核心信息"],
            "trigger_scenarios": ["baseline"],
            "must_cover": ["允许模型完成主流程", "在结尾前提出1个关键确认问题"],
            "domain": domain,
        },
        {
            "profile_id": "silent_user",
            "persona": "沉默型用户",
            "goal": "只做极简回应，测试模型能否主动推进流程而不过度重复。",
            "style": ["少字回答", "被动", "不主动提问"],
            "trigger_scenarios": ["minimal_response"],
            "must_cover": ["连续两轮只给极简反馈", "观察模型是否仍能推进主流程"],
            "domain": domain,
        },
        {
            "profile_id": "busy_user",
            "persona": "忙碌型用户",
            "goal": "尽快结束通话，观察模型是否能压缩表达并保留重点。",
            "style": ["打断", "催促", "低耐心"],
            "trigger_scenarios": ["user_says_busy_early"],
            "must_cover": ["在前两轮内表达很忙", "若模型继续说明则保持短句回应"],
            "domain": domain,
        },
        {
            "profile_id": "skeptical_user",
            "persona": "质疑型用户",
            "goal": "追问规则依据，测试模型是否按知识点稳定回答。",
            "style": ["追问细节", "怀疑规则", "要求解释原因"],
            "trigger_scenarios": ["faq_challenge"],
            "must_cover": ["至少追问2个知识点", "检查是否出现编造信息"],
            "domain": domain,
        },
        {
            "profile_id": "off_topic_user",
            "persona": "越界提问用户",
            "goal": "提出超出职责范围的问题，测试兜底回复。",
            "style": ["跳出主题", "询问非脚本内容"],
            "trigger_scenarios": ["out_of_scope_question"],
            "must_cover": ["提出1个超范围问题", "观察是否使用指定兜底策略"],
            "domain": domain,
        },
        {
            "profile_id": "random_question_user",
            "persona": "乱问问题用户",
            "goal": "在流程中穿插无关或半相关问题，测试模型的抗干扰能力。",
            "style": ["跳跃提问", "想到什么问什么", "偶尔切回正题"],
            "trigger_scenarios": ["random_interrupt_question"],
            "must_cover": ["至少打断1次主流程", "提出1个半相关问题和1个无关问题"],
            "domain": domain,
        },
    ]

    if example["constraints"].get("driving_rule"):
        profiles.append(
            {
                "profile_id": "driving_user",
                "persona": "开车中用户",
                "goal": "快速告知自己在开车，检查模型是否礼貌结束。",
                "style": ["立即打断", "强调不方便接听"],
                "trigger_scenarios": ["user_is_driving"],
                "must_cover": ["首轮或次轮明确说在开车", "若模型不停则重复一次后结束"],
                "domain": domain,
            }
        )

    if any("无法配送" in item or "不想配送" in item for item in collect_required_information(example)):
        profiles.append(
            {
                "profile_id": "refusing_user",
                "persona": "拒绝执行用户",
                "goal": "明确表示不愿执行任务，测试挽留与结束策略。",
                "style": ["消极", "拒绝", "重复表达做不到"],
                "trigger_scenarios": ["refuse_task"],
                "must_cover": ["至少两次表达拒绝", "观察模型是否先安抚再结束"],
                "domain": domain,
            }
        )

    return profiles


def build_evaluation_instruction(example: dict[str, Any]) -> dict[str, Any]:
    return {
        "instruction_id": example["instruction_id"],
        "source": example["source"],
        "domain": infer_domain(example),
        "instruction_core": {
            "role": example["role"],
            "task": example["task"],
            "opening_line": example["opening_line"],
            "required_information": collect_required_information(example),
            "call_flow": example["call_flow"],
            "knowledge_points": example["knowledge_points"],
            "constraints": example["constraints"],
        },
        "success_criteria": build_success_criteria(example),
        "failure_conditions": build_failure_conditions(example),
        "user_simulator_profiles": build_user_simulator_profiles(example),
        "raw_markdown": example["raw_markdown"],
    }


def normalize_call_flow_text(flow_text: str) -> dict[str, Any]:
    if not flow_text:
        return {"steps": [], "raw_text": ""}

    roots = parse_section_tree(flow_text)
    raw_text = clean_text(flow_text)
    top_level_steps: list[dict[str, Any]] = []

    if roots:
        for index, child in enumerate(roots, start=1):
            item = {
                "step_id": f"step_{index}",
                "title": clean_inline_text(child.title),
                "instruction": section_body(child),
                "bullet_points": split_bullets(section_body(child)),
            }
            substeps = []
            for sub_index, sub in enumerate(child.children, start=1):
                substeps.append(
                    {
                        "substep_id": f"{item['step_id']}_{sub_index}",
                        "title": clean_inline_text(sub.title),
                        "instruction": section_body(sub),
                        "bullet_points": split_bullets(section_body(sub)),
                    }
                )
            if substeps:
                item["substeps"] = substeps
            top_level_steps.append(item)
    else:
        top_level_steps = [
            {
                "step_id": f"step_{index}",
                "title": f"步骤{index}",
                "instruction": bullet,
                "bullet_points": [],
            }
            for index, bullet in enumerate(split_bullets(raw_text), start=1)
        ]

    return {
        "steps": top_level_steps,
        "raw_text": raw_text,
    }


def normalize_instruction(sample_id: Any, raw_text: str) -> dict[str, Any]:
    text = clean_text(raw_text)
    sections = extract_canonical_sections(text)

    constraints_list = split_bullets(sections.get("constraints", ""))
    knowledge_points = split_bullets(sections.get("knowledge_points", ""))
    call_flow = normalize_call_flow_text(sections.get("call_flow", ""))

    result = {
        "instruction_id": str(sample_id),
        "source": {
            "file_name": INPUT_XLSX.name,
            "sheet_name": "Sheet1",
            "row_id": sample_id,
        },
        "role": sections.get("role", ""),
        "task": sections.get("task", ""),
        "opening_line": sections.get("opening_line", ""),
        "call_flow": call_flow,
        "knowledge_points": knowledge_points,
        "constraints": {
            "items": constraints_list,
            **extract_special_rules(constraints_list),
        },
        "raw_markdown": text,
    }

    if not result["knowledge_points"]:
        result["knowledge_points"] = []

    return result


def build_schema() -> dict[str, Any]:
    return {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": "https://example.com/schemas/dialog_instruction.schema.json",
        "title": "Dialog Instruction Dataset",
        "description": "标准化后的复杂外呼任务对话指令数据集。",
        "type": "object",
        "required": [
            "dataset_name",
            "source_file",
            "schema_ref",
            "example_count",
            "examples",
        ],
        "properties": {
            "dataset_name": {"type": "string"},
            "source_file": {"type": "string"},
            "schema_ref": {"type": "string"},
            "example_count": {"type": "integer", "minimum": 0},
            "examples": {
                "type": "array",
                "items": {"$ref": "#/$defs/instruction_example"},
            },
        },
        "$defs": {
            "instruction_example": {
                "type": "object",
                "required": [
                    "instruction_id",
                    "source",
                    "role",
                    "task",
                    "opening_line",
                    "call_flow",
                    "knowledge_points",
                    "constraints",
                    "raw_markdown",
                ],
                "properties": {
                    "instruction_id": {"type": "string"},
                    "source": {
                        "type": "object",
                        "required": ["file_name", "sheet_name", "row_id"],
                        "properties": {
                            "file_name": {"type": "string"},
                            "sheet_name": {"type": "string"},
                            "row_id": {"type": ["string", "integer", "number"]},
                        },
                        "additionalProperties": False,
                    },
                    "role": {"type": "string"},
                    "task": {"type": "string"},
                    "opening_line": {"type": "string"},
                    "call_flow": {
                        "type": "object",
                        "required": ["steps", "raw_text"],
                        "properties": {
                            "raw_text": {"type": "string"},
                            "steps": {
                                "type": "array",
                                "items": {"$ref": "#/$defs/step"},
                            },
                        },
                        "additionalProperties": False,
                    },
                    "knowledge_points": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "constraints": {
                        "type": "object",
                        "required": ["items"],
                        "properties": {
                            "items": {
                                "type": "array",
                                "items": {"type": "string"},
                            },
                            "max_chars_per_turn": {"type": "integer", "minimum": 1},
                            "suggested_char_range": {
                                "type": "array",
                                "items": {"type": "integer", "minimum": 1},
                                "minItems": 2,
                                "maxItems": 2,
                            },
                            "forbidden_expressions": {
                                "type": "array",
                                "items": {"type": "string"},
                            },
                            "driving_rule": {"type": "string"},
                            "out_of_scope_reply": {"type": "string"},
                            "has_end_call_condition": {"type": "boolean"},
                        },
                        "additionalProperties": False,
                    },
                    "raw_markdown": {"type": "string"},
                },
                "additionalProperties": False,
            },
            "step": {
                "type": "object",
                "required": ["step_id", "title", "instruction", "bullet_points"],
                "properties": {
                    "step_id": {"type": "string"},
                    "title": {"type": "string"},
                    "instruction": {"type": "string"},
                    "bullet_points": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "substeps": {
                        "type": "array",
                        "items": {"$ref": "#/$defs/substep"},
                    },
                },
                "additionalProperties": False,
            },
            "substep": {
                "type": "object",
                "required": ["substep_id", "title", "instruction", "bullet_points"],
                "properties": {
                    "substep_id": {"type": "string"},
                    "title": {"type": "string"},
                    "instruction": {"type": "string"},
                    "bullet_points": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                },
                "additionalProperties": False,
            },
        },
        "additionalProperties": False,
    }


def build_examples() -> dict[str, Any]:
    workbook = load_workbook(INPUT_XLSX, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    examples = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_id, instruction_text = row[:2]
        if row_id is None or not instruction_text:
            continue
        examples.append(normalize_instruction(row_id, instruction_text))

    return {
        "dataset_name": "命题二：外呼任务对话模型指令示例",
        "source_file": str(INPUT_XLSX),
        "schema_ref": str(SCHEMA_PATH),
        "example_count": len(examples),
        "examples": examples,
    }


def build_eval_schema() -> dict[str, Any]:
    return {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": "https://example.com/schemas/dialog_instruction_eval.schema.json",
        "title": "Dialog Instruction Evaluation Dataset",
        "description": "面向用户模拟器和自动评分器的复杂外呼任务评测数据集。",
        "type": "object",
        "required": [
            "dataset_name",
            "source_file",
            "schema_ref",
            "example_count",
            "examples",
        ],
        "properties": {
            "dataset_name": {"type": "string"},
            "source_file": {"type": "string"},
            "schema_ref": {"type": "string"},
            "example_count": {"type": "integer", "minimum": 0},
            "examples": {
                "type": "array",
                "items": {"$ref": "#/$defs/eval_instruction"},
            },
        },
        "$defs": {
            "eval_instruction": {
                "type": "object",
                "required": [
                    "instruction_id",
                    "source",
                    "domain",
                    "instruction_core",
                    "success_criteria",
                    "failure_conditions",
                    "user_simulator_profiles",
                    "raw_markdown",
                ],
                "properties": {
                    "instruction_id": {"type": "string"},
                    "source": {
                        "type": "object",
                        "required": ["file_name", "sheet_name", "row_id"],
                        "properties": {
                            "file_name": {"type": "string"},
                            "sheet_name": {"type": "string"},
                            "row_id": {"type": ["string", "integer", "number"]},
                        },
                        "additionalProperties": False,
                    },
                    "domain": {"type": "string"},
                    "instruction_core": {
                        "type": "object",
                        "required": [
                            "role",
                            "task",
                            "opening_line",
                            "required_information",
                            "call_flow",
                            "knowledge_points",
                            "constraints",
                        ],
                        "properties": {
                            "role": {"type": "string"},
                            "task": {"type": "string"},
                            "opening_line": {"type": "string"},
                            "required_information": {
                                "type": "array",
                                "items": {"type": "string"},
                            },
                            "call_flow": {"$ref": "#/$defs/call_flow"},
                            "knowledge_points": {
                                "type": "array",
                                "items": {"type": "string"},
                            },
                            "constraints": {"$ref": "#/$defs/constraints"},
                        },
                        "additionalProperties": False,
                    },
                    "success_criteria": {
                        "type": "array",
                        "items": {"$ref": "#/$defs/success_criterion"},
                    },
                    "failure_conditions": {
                        "type": "array",
                        "items": {"$ref": "#/$defs/failure_condition"},
                    },
                    "user_simulator_profiles": {
                        "type": "array",
                        "items": {"$ref": "#/$defs/user_simulator_profile"},
                    },
                    "raw_markdown": {"type": "string"},
                },
                "additionalProperties": False,
            },
            "call_flow": {
                "type": "object",
                "required": ["steps", "raw_text"],
                "properties": {
                    "raw_text": {"type": "string"},
                    "steps": {
                        "type": "array",
                        "items": {"$ref": "#/$defs/step"},
                    },
                },
                "additionalProperties": False,
            },
            "constraints": {
                "type": "object",
                "required": ["items"],
                "properties": {
                    "items": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "max_chars_per_turn": {"type": "integer", "minimum": 1},
                    "suggested_char_range": {
                        "type": "array",
                        "items": {"type": "integer", "minimum": 1},
                        "minItems": 2,
                        "maxItems": 2,
                    },
                    "forbidden_expressions": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "driving_rule": {"type": "string"},
                    "out_of_scope_reply": {"type": "string"},
                    "has_end_call_condition": {"type": "boolean"},
                },
                "additionalProperties": False,
            },
            "success_criterion": {
                "type": "object",
                "required": [
                    "criterion_id",
                    "category",
                    "description",
                    "priority",
                    "evidence_type",
                    "source_fields",
                ],
                "properties": {
                    "criterion_id": {"type": "string"},
                    "category": {"type": "string"},
                    "description": {"type": "string"},
                    "priority": {"type": "string"},
                    "evidence_type": {"type": "string"},
                    "source_fields": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                },
                "additionalProperties": False,
            },
            "failure_condition": {
                "type": "object",
                "required": [
                    "condition_id",
                    "severity",
                    "description",
                    "trigger_type",
                    "source_fields",
                ],
                "properties": {
                    "condition_id": {"type": "string"},
                    "severity": {"type": "string"},
                    "description": {"type": "string"},
                    "trigger_type": {"type": "string"},
                    "source_fields": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                },
                "additionalProperties": False,
            },
            "user_simulator_profile": {
                "type": "object",
                "required": [
                    "profile_id",
                    "persona",
                    "goal",
                    "style",
                    "trigger_scenarios",
                    "must_cover",
                    "domain",
                ],
                "properties": {
                    "profile_id": {"type": "string"},
                    "persona": {"type": "string"},
                    "goal": {"type": "string"},
                    "style": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "trigger_scenarios": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "must_cover": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "domain": {"type": "string"},
                },
                "additionalProperties": False,
            },
            "step": {
                "type": "object",
                "required": ["step_id", "title", "instruction", "bullet_points"],
                "properties": {
                    "step_id": {"type": "string"},
                    "title": {"type": "string"},
                    "instruction": {"type": "string"},
                    "bullet_points": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "substeps": {
                        "type": "array",
                        "items": {"$ref": "#/$defs/substep"},
                    },
                },
                "additionalProperties": False,
            },
            "substep": {
                "type": "object",
                "required": ["substep_id", "title", "instruction", "bullet_points"],
                "properties": {
                    "substep_id": {"type": "string"},
                    "title": {"type": "string"},
                    "instruction": {"type": "string"},
                    "bullet_points": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                },
                "additionalProperties": False,
            },
        },
        "additionalProperties": False,
    }


def build_eval_examples(base_examples: dict[str, Any]) -> dict[str, Any]:
    eval_examples = [
        build_evaluation_instruction(example)
        for example in base_examples["examples"]
    ]
    return {
        "dataset_name": f"{base_examples['dataset_name']}（评测版）",
        "source_file": base_examples["source_file"],
        "schema_ref": str(EVAL_SCHEMA_PATH),
        "example_count": len(eval_examples),
        "examples": eval_examples,
    }


def main() -> None:
    ensure_dirs()

    schema = build_schema()
    examples = build_examples()
    eval_schema = build_eval_schema()
    eval_examples = build_eval_examples(examples)

    SCHEMA_PATH.write_text(
        json.dumps(schema, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    OUTPUT_PATH.write_text(
        json.dumps(examples, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    EVAL_SCHEMA_PATH.write_text(
        json.dumps(eval_schema, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    EVAL_OUTPUT_PATH.write_text(
        json.dumps(eval_examples, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Wrote schema to: {SCHEMA_PATH}")
    print(f"Wrote examples to: {OUTPUT_PATH}")
    print(f"Wrote eval schema to: {EVAL_SCHEMA_PATH}")
    print(f"Wrote eval examples to: {EVAL_OUTPUT_PATH}")
    print(f"Extracted examples: {examples['example_count']}")


if __name__ == "__main__":
    main()
